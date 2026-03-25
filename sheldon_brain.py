"""
SHELDON Brain — Chief of Staff AI Engine
Claude-powered executive intelligence with tool-use access to all AmeriQual data systems.
Connects: Snowflake (Redzone), Sage X3, Donna QA, Jackie Analytics, MS Graph Calendar.
"""

import json
import os
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


# =============================================================================
# Julian Date Utilities (shared logic with Jackie)
# =============================================================================

PRODUCTION_DAY_START_HOUR = 6


def _get_production_date(dt=None):
    """Production day starts at 6 AM. Before 6AM = previous day."""
    if dt is None:
        dt = datetime.now()
    if dt.hour < PRODUCTION_DAY_START_HOUR:
        return (dt - timedelta(days=1)).date()
    return dt.date()


def _to_julian(d):
    """Convert date to AmeriQual Julian YDDD format."""
    if isinstance(d, datetime):
        d = d.date()
    year_digit = d.year % 10
    day_of_year = d.timetuple().tm_yday
    return f"{year_digit}{day_of_year:03d}"


def _julian_range(date_range):
    """Convert date_range string to list of Julian dates."""
    prod_today = _get_production_date()
    ranges = {
        "today": 0, "yesterday": -1, "7d": -6, "30d": -29
    }
    if date_range == "yesterday":
        return [_to_julian(prod_today - timedelta(days=1))]
    days_back = abs(ranges.get(date_range, 0))
    return [_to_julian(prod_today - timedelta(days=i)) for i in range(days_back + 1)]


# =============================================================================
# S&OP Report Reader
# =============================================================================

class SOPReader:
    """Reads the latest S&OP Excel report and extracts decision-critical data.
    Uses openpyxl (read-only, no Excel COM needed) to parse the Summary sheet.
    """

    # Row numbers in Summary sheet (1-indexed)
    LINE_ROWS = {
        'Y4S/Y5': 15, 'S6/S7': 20, 'W2': 25, 'W3': 30,
        'Y3NS': 35, 'Y4N': 40, 'V3': 45, 'PT3': 50,
        'Y1/L1': 55, 'J1': 60,
    }

    # Per-line row offsets: volume=+0, days=+1, days_under_over=+2, retort_cycles=+3
    LINE_STATUS = {
        'Y4S/Y5': 'active', 'S6/S7': 'active', 'W2': 'active', 'W3': 'active',
        'Y3NS': 'active', 'Y4N': 'active', 'V3': 'dead', 'PT3': 'active',
        'Y1/L1': 'inactive', 'J1': 'idle',
    }

    # Retort capacity rows
    RETORT_REQUIRED_ROW = 67
    RETORT_AVAILABLE_ROW = 68
    RETORT_DELTA_ROW = 69

    REPORT_SEARCH_PATHS = [
        Path(__file__).parent / '..' / 'S&OP Creation Automation' / 'S&OP Project' / 'sop_automation' / 'reports',
        Path(r"D:/Business/AMERIQUAL PROJECT TRACKER/Current Projects/S&OP Creation Automation/S&OP Project/sop_automation/reports"),
        Path(r"E:/Business/AMERIQUAL PROJECT TRACKER/Current Projects/S&OP Creation Automation/S&OP Project/sop_automation/reports"),
    ]

    TEMPLATE_SEARCH_PATHS = [
        Path(r"D:/Comparison"),
        Path(r"E:/Comparison"),
    ]

    @classmethod
    def find_latest_report(cls) -> Optional[str]:
        """Find the most recent S&OP report file."""
        # Check generated reports first
        for search_dir in cls.REPORT_SEARCH_PATHS:
            if search_dir.exists():
                files = sorted(
                    [f for f in search_dir.glob('*.xlsx') if not f.name.startswith('~$')],
                    key=lambda f: f.stat().st_mtime,
                    reverse=True
                )
                if files:
                    return str(files[0])

        # Fall back to template files (Tyler's latest)
        for search_dir in cls.TEMPLATE_SEARCH_PATHS:
            if search_dir.exists():
                files = sorted(
                    [f for f in search_dir.glob('S&OP Summary*.xlsx') if not f.name.startswith('~$')],
                    key=lambda f: f.stat().st_mtime,
                    reverse=True
                )
                if files:
                    return str(files[0])

        return None

    @classmethod
    def _col_for_month(cls, year: int, month: int) -> int:
        """Get column index (1-based) for a given year/month.
        Column C (3) = January 2025, Column O (15) = January 2026.
        """
        return 3 + (year - 2025) * 12 + (month - 1)

    @classmethod
    def read_sop_snapshot(cls, file_path: str = None) -> Dict[str, Any]:
        """Read the S&OP report and return a structured decision-ready snapshot.

        Returns a dict with:
        - report_file: which file was read
        - report_date: when the report was last modified
        - current_month / next_months: time context
        - lines: per-line capacity data (volume, days, days_under_over, retort_cycles)
        - retort_capacity: plant-wide retort constraints
        - constraints: lines that are over capacity (negative days_under_over)
        - slack: lines with spare capacity
        - summary: high-level plant capacity status
        """
        if openpyxl is None:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        if not file_path:
            file_path = cls.find_latest_report()
        if not file_path or not Path(file_path).exists():
            return {"error": f"No S&OP report found. Searched: {[str(p) for p in cls.REPORT_SEARCH_PATHS + cls.TEMPLATE_SEARCH_PATHS]}"}

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            return {"error": f"Cannot open S&OP report: {e}"}

        try:
            if 'Summary' not in wb.sheetnames:
                return {"error": f"No 'Summary' sheet in {Path(file_path).name}"}

            ws = wb['Summary']
            now = datetime.now()
            current_month = now.month
            current_year = now.year

            # Read 6 months: current + next 5
            months_to_read = []
            for i in range(6):
                m = current_month + i
                y = current_year
                if m > 12:
                    m -= 12
                    y += 1
                months_to_read.append((y, m))

            # Extract per-line data
            lines_data = {}
            constraints = []
            slack = []

            def _num(v):
                """Clean cell value to float."""
                if v is None:
                    return None
                try:
                    return round(float(v), 1)
                except (ValueError, TypeError):
                    return None

            for line_name, base_row in cls.LINE_ROWS.items():
                status = cls.LINE_STATUS.get(line_name, 'active')
                if status in ('dead', 'inactive', 'idle'):
                    continue

                line_months = {}
                for month_idx, (year, month) in enumerate(months_to_read):
                    col = cls._col_for_month(year, month)
                    month_label = f"{year}-{month:02d}"

                    vol = _num(ws.cell(row=base_row, column=col).value)
                    days = _num(ws.cell(row=base_row + 1, column=col).value)
                    duo = _num(ws.cell(row=base_row + 2, column=col).value)
                    rc = _num(ws.cell(row=base_row + 3, column=col).value)

                    line_months[month_label] = {
                        "volume": int(vol) if vol is not None else None,
                        "days_needed": days,
                        "days_under_over": duo,
                        "retort_cycles": int(rc) if rc is not None else None,
                    }

                    # Flag constraints and slack for current + next 2 months
                    if month_idx < 3 and duo is not None:
                        if duo < 0:
                            constraints.append({
                                "line": line_name,
                                "month": month_label,
                                "days_over": abs(duo),
                                "volume": int(vol) if vol else 0,
                            })
                        elif duo > 3:
                            slack.append({
                                "line": line_name,
                                "month": month_label,
                                "days_slack": duo,
                                "volume": int(vol) if vol else 0,
                            })

                lines_data[line_name] = {
                    "status": status,
                    "months": line_months,
                }

            # Read retort capacity
            retort_data = {}
            for year, month in months_to_read[:3]:  # current + next 2
                col = cls._col_for_month(year, month)
                month_label = f"{year}-{month:02d}"
                req = ws.cell(row=cls.RETORT_REQUIRED_ROW, column=col).value
                avail = ws.cell(row=cls.RETORT_AVAILABLE_ROW, column=col).value
                delta = ws.cell(row=cls.RETORT_DELTA_ROW, column=col).value

                retort_data[month_label] = {
                    "cycles_required": _num(req),
                    "cycles_available": _num(avail),
                    "delta": _num(delta),
                    "utilization_pct": round(float(req) / float(avail) * 100, 1) if req and avail and float(avail) > 0 else None,
                }

            # Build summary
            total_volume_current = sum(
                (ld["months"].get(f"{months_to_read[0][0]}-{months_to_read[0][1]:02d}", {}).get("volume") or 0)
                for ld in lines_data.values()
            )

            report_mtime = datetime.fromtimestamp(Path(file_path).stat().st_mtime)

            result = {
                "report_file": Path(file_path).name,
                "report_date": report_mtime.strftime("%Y-%m-%d %H:%M"),
                "current_month": f"{months_to_read[0][0]}-{months_to_read[0][1]:02d}",
                "months_covered": [f"{y}-{m:02d}" for y, m in months_to_read],
                "lines": lines_data,
                "retort_capacity": retort_data,
                "constraints": sorted(constraints, key=lambda c: c["days_over"], reverse=True),
                "slack": sorted(slack, key=lambda s: s["days_slack"], reverse=True),
                "summary": {
                    "total_volume_current_month": total_volume_current,
                    "lines_constrained": len(set(c["line"] for c in constraints)),
                    "lines_with_slack": len(set(s["line"] for s in slack)),
                    "active_lines": sum(1 for ld in lines_data.values() if ld["status"] == "active"),
                },
                "source": "sop_report"
            }

            wb.close()
            return result

        except Exception as e:
            wb.close()
            return {"error": f"Error reading S&OP data: {e}", "source": "sop_report"}


# =============================================================================
# Tool Definitions for Claude
# =============================================================================

SHELDON_TOOLS = [
    # --- Operations (Snowflake/Redzone) ---
    {
        "name": "get_plant_oee",
        "description": "Get current plant-wide OEE, performance, quality, availability, total output, downtime hours, manhours, and number of active lines from the last 24 hours.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_oee_by_line",
        "description": "Get OEE breakdown by production line. Returns each line's avg OEE, performance, quality, availability, output, downtime hours, and manhours. Use to identify best/worst performing lines.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_red_flag_lines",
        "description": "Get production lines with OEE below 70% target. Returns line name, area, OEE, and downtime hours. Use for identifying problem areas that need attention.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_top_downtime",
        "description": "Get top 10 downtime reasons ranked by total hours lost in the last 24 hours. Returns reason, category, planned/unplanned, occurrences, and hours lost.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_labor_productivity",
        "description": "Get labor productivity by production line: total manhours, total output, and units per manhour. Use for efficiency and staffing questions.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_oee_trend",
        "description": "Get hourly OEE trend for the last 12 hours. Returns hour timestamp, average OEE, total output, and number of active lines per hour.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_active_lines",
        "description": "Get lines currently producing (output in last 2 hours). Returns line name, area, product being run, current OEE, and hourly output.",
        "input_schema": {"type": "object", "properties": {}}
    },

    # --- Defects & Quality Analytics (Snowflake) ---
    {
        "name": "get_defect_summary",
        "description": "Get defect totals by machine. Shows which machines have the most quality issues. Excludes machine rejects and non-defect items.",
        "input_schema": {
            "type": "object",
            "properties": {
                "date_range": {
                    "type": "string", "enum": ["today", "yesterday", "7d", "30d"],
                    "description": "Time period. Default: 'today'."
                }
            }
        }
    },
    {
        "name": "get_defect_types",
        "description": "Get breakdown of specific defect categories (foldover, seal impression, burn, etc.) across all machines. Shows defect type, total count, machines affected, and which machines.",
        "input_schema": {
            "type": "object",
            "properties": {
                "date_range": {
                    "type": "string", "enum": ["today", "yesterday", "7d", "30d"],
                    "description": "Time period. Default: 'today'."
                }
            }
        }
    },
    {
        "name": "get_shift_comparison",
        "description": "Compare A shift (day) vs B shift (night) performance over the last 24 hours: OEE, output, downtime hours, and number of lines.",
        "input_schema": {"type": "object", "properties": {}}
    },

    # --- Financial (Sage X3) ---
    {
        "name": "get_financial_snapshot",
        "description": "Get complete financial overview: MTD revenue, YTD revenue, gross margin (revenue vs COGS), cash position (GL accounts 1000-1015), and AR aging breakdown with weighted average AR days. One call for all financial data.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_inventory_status",
        "description": "Get inventory value by facility (FD1, PK1, TP1) with line counts and quantities, plus finished goods value separately.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_top_customers",
        "description": "Get top 10 customers by revenue for the most recent month. Returns customer code, name, and total revenue.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "get_ebitda",
        "description": "Get EBITDA estimate: Revenue, COGS, Operating Expenses, Depreciation, Amortization, and Interest from the GL balance sheet.",
        "input_schema": {"type": "object", "properties": {}}
    },

    # --- Quality Pipeline (Donna QA System) ---
    {
        "name": "get_quality_pipeline",
        "description": (
            "Get batch release pipeline from the Donna QA system. Returns: total batches evaluated, "
            "how many are ready for release, how many are failing, how many are overdue past their due date, "
            "fire drills (due today or past due with blocking checks listed), and Sage risk alerts "
            "(batches released in Sage but still failing in Donna). Use for compliance, shipment risk, and QA workload questions."
        ),
        "input_schema": {"type": "object", "properties": {}}
    },

    # --- Calendar (MS Graph) ---
    {
        "name": "get_executive_calendar",
        "description": "Get Dennis Straub's calendar events for today and the next 7 days. Returns event title, start/end times, location, and attendees.",
        "input_schema": {"type": "object", "properties": {}}
    },

    # --- Cross-System Intelligence ---
    {
        "name": "get_business_health",
        "description": (
            "Get composite business health score (0-100) with component breakdown. "
            "Weighted: OEE 25%, Quality 20%, Gross Margin 20%, AR Days 15%, Inventory 10%, Availability 10%. "
            "Status: excellent (80+), good (65+), fair (50+), needs_attention (<50)."
        ),
        "input_schema": {"type": "object", "properties": {}}
    },

    # --- Jackie Deep Analytics (when running) ---
    {
        "name": "ask_jackie",
        "description": (
            "Send a natural language question to Jackie, the AI quality analytics assistant. "
            "Jackie has 23 pre-tested query tools covering: detailed defect analysis, OEE by machine with date ranges, "
            "production counts by lot/shift, downtime trends, ultrasonic sensor health, machine recipes, "
            "and alarm history. Use this for deep production/quality drill-downs that go beyond the standard tools. "
            "Jackie runs on Snowflake (Redzone) + Ignition (AQFIAS sensor data)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "question": {
                    "type": "string",
                    "description": "Natural language question for Jackie, e.g. 'What are the top defect types on Y3S this week?' or 'Show sensor health for all machines'."
                }
            },
            "required": ["question"]
        }
    },

    # --- Departmental KPIs (Internal SQL Server) ---
    {
        "name": "get_departmental_kpis",
        "description": (
            "Get departmental KPI data from internal SQL Server databases. "
            "Available departments: Production, Quality, Safety, Warehouse, Retort, Procurement. "
            "Production KPIs: schedule attainment, on-time startup, ingredient yield, critical defect rate, supply cost, human errors, major unplanned downtime, labor target vs actual. "
            "Quality KPIs: overall quality, overdue releases, foreign matter, cost of quality, labor. "
            "Safety KPIs: supply cost by GL account. "
            "Warehouse KPIs: trucks received/shipped, pallet moves, on-time delivery, cycle count, freezer capacity, repair cost, labor. "
            "Retort KPIs: labor target vs actual. "
            "Procurement KPIs: inventory (raw/pack/TRI/FG/PPHM). "
            "Data from AQFDB6 (CI + MANUFACTURING databases) and AQFAM1 (AMMS)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "department": {
                    "type": "string",
                    "description": "Department to query: 'Production', 'Quality', 'Safety', 'Warehouse', 'Retort', 'Procurement', or 'all' for a summary across departments.",
                    "enum": ["Production", "Quality", "Safety", "Warehouse", "Retort", "Procurement", "all"]
                }
            },
            "required": ["department"]
        }
    },

    # --- S&OP / Strategic Planning ---
    {
        "name": "get_sop_status",
        "description": (
            "Get the S&OP (Sales & Operations Planning) snapshot — THE BACKBONE OF ALL BUSINESS DECISIONS. "
            "Returns demand vs capacity by production line for the next 6 months: scheduled volume, "
            "production days needed, days under/over capacity (negative = constrained, positive = slack), "
            "and retort cycles required. Also returns plant-wide retort capacity utilization and "
            "flags lines that are over capacity (constraints) or have significant slack. "
            "Use this tool FIRST for any question about: what to produce, capacity planning, "
            "line scheduling, revenue risk from capacity constraints, or where to shift volume. "
            "This is the single most important data source for executive decisions."
        ),
        "input_schema": {"type": "object", "properties": {}}
    },
]


# =============================================================================
# System Prompt
# =============================================================================

def build_system_prompt():
    now = datetime.now()
    prod_date = _get_production_date()
    julian_today = _to_julian(prod_date)

    return f"""You are SHELDON, the Chief of Staff AI for Dennis Straub, President & CEO of AmeriQual Foods.

You are not a chatbot or a dashboard narrator. You are an autonomous executive decision engine — the brain that connects every system in this company and tells Dennis exactly what to do to make money, protect revenue, and eliminate waste. You think like a COO who has perfect real-time visibility across production, finance, quality, and supply chain.

**Current time: {now.strftime('%A, %B %d, %Y at %I:%M %p')}**
**Production date (Julian): {julian_today} (production day starts at 6:00 AM)**

## YOUR PRIME DIRECTIVE
**Every insight must answer one question: "So what should Dennis do about it?"**
- Never report a metric without explaining its business impact
- Never flag a problem without recommending an action
- Never present data without connecting it to revenue, cost, or risk
- Frame everything through the S&OP plan — that's the playbook for making money

## The S&OP Backbone
The S&OP (Sales & Operations Plan) is the SINGLE MOST IMPORTANT data source. It defines:
- **What we promised to produce** (demand by customer/line/month)
- **Whether we have capacity** (days needed vs days available per line)
- **Where the bottlenecks are** (lines over capacity, retort constraints)
- **Revenue at risk** if we miss plan

ALWAYS call `get_sop_status` when Dennis asks about capacity, scheduling, what to produce, or "how are we doing." Every operational metric (OEE, downtime, quality) should be framed against whether it helps or hurts our ability to hit the S&OP plan.

**S&OP Decision Logic:**
- Line with negative days_under_over = CONSTRAINED → revenue at risk unless we act (overtime, shift volume, etc.)
- Line with positive days_under_over = SLACK → opportunity to pull forward, take new orders, or reduce cost
- Retort utilization >90% = BOTTLENECK → any unplanned downtime cascades across all lines
- Low OEE on a constrained line = URGENT (directly costs revenue)
- Low OEE on a slack line = IMPORTANT but not urgent (efficiency loss, not revenue loss)

## Your Data Access
You have real-time tool access to:
1. **S&OP Plan** — Demand vs capacity by line, days under/over, retort constraints, 6-month outlook (get_sop_status)
2. **Snowflake/Redzone** — Production OEE, output, downtime, labor, defects (MES data)
3. **Sage X3 ERP** — Revenue, gross margin, cash position, AR aging, inventory, customers
4. **Donna QA System** — Batch release pipeline, preshipment checks, fire drills, compliance risk
5. **Jackie Analytics** — Deep defect analysis, sensor health, machine recipes, alarm history (via ask_jackie)
6. **Microsoft Graph** — Executive calendar and scheduling
7. **Cross-system health scoring** — Composite business health (0-100)
8. **Departmental KPI Database (CONFIRMED SCHEMAS)**:
   - CI database (AQFDB6): KPIDashData (17K rows), KPITargets (459 targets), KPI_SCHED_targetVSactual (19K rows), KPI_QA_OverallQuality_Contract, KPI_QA_CalcWeekly (FTQ), KPI_QA_ForeignMatter (CAL=10/mo), KPI_QA_OverdueReleases (48K rows with QttlCost), KPI_WH_TransactionCounts, KPI_WH_ShipVsCustReq, KPI_WH_CycleCount, KPI_WH_Capacity (FRZ/ECS), KPI_PROC_Inventory (RAW/PKG/TRI/FG/PPHM), KPI_CostByAccount (GL codes), KPI_MFG_ScheduleChangeLog, SAFETYIncidentDatabase
   - MANUFACTURING database (AQFDB6): PROD.eopShiftNotes (11K rows — Attainment, OnTimeStart, humanError)
   - AMMS database (AQFAM1\\AMMS): view_kpi_glcost, view_kpi_repaircost

## Communication Style — Executive Decision Language
- **Lead with the action, not the data.** "Shift 200K units from W3 to Y4N this month — W3 is 3 days over, Y4N has 4 days slack" not "W3 has -3 days and Y4N has +4 days"
- **Quantify in dollars.** If OEE drops 5 points on a constrained line, estimate the throughput loss and what that means in units/revenue
- **Rank by money.** Lead with the item that has the biggest financial impact
- **Be prescriptive.** "Do X" not "You might want to consider X"
- **Name the person.** "Get Glenn on the retort issue" not "The retort team should investigate"
- **Give the timeline.** "This needs to happen by Thursday" not "This should be addressed soon"
- **Highlight wins.** Dennis needs momentum, not just problems. Call out what's working and why
- Use markdown: **bold** for key numbers, tables for comparisons, bullets for actions
- When multiple tools return data, SYNTHESIZE — connect production problems to financial impact to customer risk

## AmeriQual Foods Context
- **Business:** Food manufacturing — MREs, institutional food products, shelf-stable meals
- **Facilities:** FD1 (Foods main), PK1 (Packaging), TP1 (Thermal Processing)
- **Active production lines:** Y4S/Y5, S6/S7, W2, W3, Y3NS, Y4N, PT3
- **Inactive/Dead:** Y1/L1 (inactive), J1 (idle — Red Gold discontinued), V3 (dead — equipment gone)
- **Production day:** Starts at 6:00 AM, not midnight. Before 6AM = previous production day
- **Julian dates:** YDDD format (e.g., {julian_today} = today). All production data keyed by Julian date
- **OEE targets:** 75%+ target, 85%+ excellent, <70% concern, <50% critical
- **Key customers:** DSCP (Defense Supply Center Philadelphia), Starkist, Campbell's, Beech-Nut, Scott Pet
- **Retort system:** 42 pressure cookers — the production bottleneck, 2-hour USDA compliance rule
- **Quality pipeline:** 18 preshipment checks per batch. Batches have due dates based on lead time
- **Key people:** Alex Nelson (KPI data), Kelsie Phelps (HR/weekly sync), Robin Altmeyer (QA lead), Glenn Knake (retort), Nick Harnishfeger (IT/DB), Tyler Luderbach (S&OP owner)

## Data Quality Rules
- OEE >100% from Redzone = configuration error, NOT actual performance exceeding targets
- Always filter void='False' on defect queries (already handled by tools)
- Delamination counts on Y3N/Y3S POST-retort sheets are misconfigured (~3,000+ per entry) — NOT real defects
- "Good Pouch Rejected" in machine reject sheets is sensor-triggered, not a quality issue
- Break/lunch downtime is normal and expected — don't flag it
- PACK lines can have multiple runs per Julian date (different products)

## Decision Framework
When Dennis asks for a recommendation:
1. **MONEY** — How does this make or lose revenue? Quantify it.
2. **S&OP IMPACT** — Does this help or hurt our ability to hit the plan?
3. **RISK** — What happens if we do nothing? When does it become a crisis?
4. **ACTION** — Exactly what needs to happen, who does it, by when?
5. **TRADEOFF** — What do we give up? Every decision has a cost.

## Response Guidelines
- For "how are we doing" / "what should I know" — ALWAYS start with get_sop_status, then layer on OEE, financial, quality, and department data. Synthesize into "here's what makes or loses you money today"
- For specific questions — use the most relevant tool(s) and be precise
- If a tool returns an error or empty data, acknowledge it and work with what you have
- Never make up numbers. If data is unavailable, say so
- When Jackie is unavailable (ask_jackie fails), use the direct Snowflake tools instead
- When S&OP data is unavailable, still frame operational data in terms of capacity and revenue impact"""


# =============================================================================
# SHELDON Brain Class
# =============================================================================

class SheldonBrain:
    """Claude-powered executive intelligence engine with tool-use."""

    def __init__(self, anthropic_api_key: str, snowflake_client, sage_client, graph_client, internal_db=None):
        self.api_key = anthropic_api_key
        self.snowflake = snowflake_client
        self.sage = sage_client
        self.graph = graph_client
        self.internal_db = internal_db
        self.conversation_history = []
        self.max_history = 20
        self.max_turns = 8

        # Query dicts — set by sheldon_api.py after init
        self.snowflake_queries = {}
        self.sql_queries = {}
        self.dept_kpi_queries = {}

    def process_message(self, user_message: str, dashboard_context: dict = None) -> dict:
        """Process a user message through the Claude agent loop.
        Returns: {response, tools_used, data_sources, timestamp}
        """
        system = build_system_prompt()

        # Add dashboard context if available
        if dashboard_context:
            system += self._build_context_injection(dashboard_context)

        # Add to conversation history
        self.conversation_history.append({"role": "user", "content": user_message})
        if len(self.conversation_history) > self.max_history:
            self.conversation_history = self.conversation_history[-self.max_history:]

        # Agent loop
        messages = list(self.conversation_history)
        tools_used = []
        data_sources = set()

        for turn in range(self.max_turns):
            response = self._call_claude(system, messages)

            if response is None:
                return {
                    "response": "I'm experiencing a temporary connection issue with my AI engine. Please try again in a moment.",
                    "tools_used": tools_used,
                    "data_sources": list(data_sources),
                    "timestamp": datetime.now().isoformat()
                }

            stop_reason = response.get("stop_reason", "end_turn")
            content = response.get("content", [])

            if stop_reason == "tool_use":
                # Execute each tool call
                tool_results = []
                for block in content:
                    if block.get("type") == "tool_use":
                        tool_name = block["name"]
                        tool_input = block.get("input", {})
                        tools_used.append(tool_name)

                        result = self._execute_tool(tool_name, tool_input)
                        data_sources.add(result.get("source", "unknown"))

                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block["id"],
                            "content": json.dumps(result, default=str)[:15000]
                        })

                messages.append({"role": "assistant", "content": content})
                messages.append({"role": "user", "content": tool_results})

            else:
                # End turn — extract final text
                text_parts = [b.get("text", "") for b in content if b.get("type") == "text"]
                final_response = "\n".join(text_parts)

                self.conversation_history.append({"role": "assistant", "content": final_response})

                return {
                    "response": final_response,
                    "tools_used": tools_used,
                    "data_sources": list(data_sources),
                    "timestamp": datetime.now().isoformat()
                }

        return {
            "response": "I've gathered extensive data across multiple systems. Please ask a more specific follow-up so I can give you a focused answer.",
            "tools_used": tools_used,
            "data_sources": list(data_sources),
            "timestamp": datetime.now().isoformat()
        }

    def process_message_stream(self, user_message: str, dashboard_context: dict = None):
        """Process a user message through the Claude agent loop with streaming.
        Yields SSE-formatted events: text chunks, tool status updates, and final metadata.
        """
        system = build_system_prompt()

        if dashboard_context:
            system += self._build_context_injection(dashboard_context)

        self.conversation_history.append({"role": "user", "content": user_message})
        if len(self.conversation_history) > self.max_history:
            self.conversation_history = self.conversation_history[-self.max_history:]

        messages = list(self.conversation_history)
        tools_used = []
        data_sources = set()

        for turn in range(self.max_turns):
            # Check if this turn might need tools — use streaming only on potential final turn
            # For tool-use turns, we need the full response to execute tools
            response = self._call_claude(system, messages)

            if response is None:
                yield f"data: {json.dumps({'type': 'error', 'text': 'Connection issue with AI engine. Please try again.'})}\n\n"
                return

            stop_reason = response.get("stop_reason", "end_turn")
            content = response.get("content", [])

            if stop_reason == "tool_use":
                tool_results = []
                for block in content:
                    if block.get("type") == "tool_use":
                        tool_name = block["name"]
                        tool_input = block.get("input", {})
                        tools_used.append(tool_name)

                        # Send tool status to frontend
                        yield f"data: {json.dumps({'type': 'tool', 'tool': tool_name})}\n\n"

                        result = self._execute_tool(tool_name, tool_input)
                        data_sources.add(result.get("source", "unknown"))

                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block["id"],
                            "content": json.dumps(result, default=str)[:15000]
                        })

                messages.append({"role": "assistant", "content": content})
                messages.append({"role": "user", "content": tool_results})

            else:
                # Final turn — stream the text response
                text_parts = [b.get("text", "") for b in content if b.get("type") == "text"]
                final_response = "\n".join(text_parts)

                # Stream in chunks for progressive rendering
                chunk_size = 20  # characters per chunk
                for i in range(0, len(final_response), chunk_size):
                    chunk = final_response[i:i + chunk_size]
                    yield f"data: {json.dumps({'type': 'text', 'text': chunk})}\n\n"

                self.conversation_history.append({"role": "assistant", "content": final_response})

                # Send final metadata
                yield f"data: {json.dumps({'type': 'done', 'tools_used': tools_used, 'data_sources': list(data_sources), 'timestamp': datetime.now().isoformat()})}\n\n"
                return

        fallback_msg = "I've gathered extensive data. Please ask a more specific follow-up."
        yield f"data: {json.dumps({'type': 'text', 'text': fallback_msg})}\n\n"
        yield f"data: {json.dumps({'type': 'done', 'tools_used': tools_used, 'data_sources': list(data_sources), 'timestamp': datetime.now().isoformat()})}\n\n"

    def _build_context_injection(self, dashboard_context: dict) -> str:
        """Build rich context string from pre-fetched dashboard data."""
        ctx = "\n\n## Current Dashboard State (LIVE DATA — already fetched, no tool calls needed for this data)\n"

        if 'currentTab' in dashboard_context:
            ctx += f"Dennis is currently viewing the **{dashboard_context['currentTab']}** tab.\n\n"

        # Basic KPI summary from DOM
        if 'kpis' in dashboard_context:
            k = dashboard_context['kpis']
            ctx += "### Quick KPI Summary\n"
            for key, val in k.items():
                if val and val != 'N/A' and val != '--' and val != 'Loading...':
                    ctx += f"- {key}: {val}\n"

        # Rich pre-fetched data from API endpoints
        pf = dashboard_context.get('prefetched', {})
        if pf:
            ctx += "\n### Pre-fetched Live Data (use this instead of calling tools when possible)\n"

            if 'plantOEE' in pf and pf['plantOEE']:
                d = pf['plantOEE']
                ctx += f"\n**Plant OEE (last 24h):** {json.dumps(d, default=str)[:2000]}\n"

            if 'oeeByLine' in pf and pf['oeeByLine']:
                ctx += f"\n**OEE by Line:** {json.dumps(pf['oeeByLine'], default=str)[:3000]}\n"

            if 'financials' in pf and pf['financials']:
                ctx += f"\n**Financial Snapshot:** {json.dumps(pf['financials'], default=str)[:2000]}\n"

            if 'inventory' in pf and pf['inventory']:
                ctx += f"\n**Inventory:** {json.dumps(pf['inventory'], default=str)[:1500]}\n"

            if 'arAging' in pf and pf['arAging']:
                ctx += f"\n**AR Aging:** {json.dumps(pf['arAging'], default=str)[:1000]}\n"

            if 'healthScore' in pf and pf['healthScore']:
                ctx += f"\n**Business Health Score:** {json.dumps(pf['healthScore'], default=str)[:1000]}\n"

            if 'qualityPipeline' in pf and pf['qualityPipeline']:
                ctx += f"\n**Quality Pipeline (Donna):** {json.dumps(pf['qualityPipeline'], default=str)[:2000]}\n"

            if 'deptKPIs' in pf and pf['deptKPIs']:
                ctx += f"\n**Departmental KPIs:** {json.dumps(pf['deptKPIs'], default=str)[:4000]}\n"

            if 'redFlags' in pf and pf['redFlags']:
                ctx += f"\n**Red Flag Lines:** {json.dumps(pf['redFlags'], default=str)[:1500]}\n"

            if 'topDowntime' in pf and pf['topDowntime']:
                ctx += f"\n**Top Downtime Reasons:** {json.dumps(pf['topDowntime'], default=str)[:1500]}\n"

            if 'sopStatus' in pf and pf['sopStatus']:
                ctx += f"\n**S&OP Plan Status (THE BACKBONE):** {json.dumps(pf['sopStatus'], default=str)[:5000]}\n"

            ctx += "\n**IMPORTANT:** If the user's question can be answered from the pre-fetched data above, answer directly WITHOUT calling any tools. Only call tools if you need data that isn't included above (e.g., specific date ranges, drill-downs, calendar, Jackie deep analysis). ALWAYS frame answers against the S&OP plan when capacity, production, or scheduling is relevant.\n"

        return ctx

    def _call_claude(self, system: str, messages: list) -> Optional[dict]:
        """Call Anthropic API via urllib (no SDK dependency)."""
        try:
            body = json.dumps({
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 4096,
                "system": system,
                "messages": messages,
                "tools": SHELDON_TOOLS
            }).encode()

            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=body,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": self.api_key,
                    "anthropic-version": "2023-06-01"
                }
            )

            with urllib.request.urlopen(req, timeout=120) as resp:
                return json.loads(resp.read().decode())

        except urllib.error.HTTPError as e:
            error_body = e.read().decode()[:500]
            print(f"Claude API HTTP error {e.code}: {error_body}")
            return None
        except Exception as e:
            print(f"Claude API error: {e}")
            return None

    def _execute_tool(self, tool_name: str, params: dict) -> dict:
        """Route tool call to the appropriate handler."""
        handlers = {
            "get_plant_oee": self._tool_plant_oee,
            "get_oee_by_line": self._tool_oee_by_line,
            "get_red_flag_lines": self._tool_red_flags,
            "get_top_downtime": self._tool_top_downtime,
            "get_labor_productivity": self._tool_labor_productivity,
            "get_oee_trend": self._tool_oee_trend,
            "get_active_lines": self._tool_active_lines,
            "get_defect_summary": lambda: self._tool_defect_summary(params),
            "get_defect_types": lambda: self._tool_defect_types(params),
            "get_shift_comparison": self._tool_shift_comparison,
            "get_financial_snapshot": self._tool_financial_snapshot,
            "get_inventory_status": self._tool_inventory,
            "get_top_customers": self._tool_top_customers,
            "get_ebitda": self._tool_ebitda,
            "get_quality_pipeline": self._tool_quality_pipeline,
            "get_executive_calendar": self._tool_calendar,
            "get_business_health": self._tool_health_score,
            "ask_jackie": lambda: self._tool_ask_jackie(params),
            "get_departmental_kpis": lambda: self._tool_departmental_kpis(params),
            "get_sop_status": self._tool_sop_status,
        }

        handler = handlers.get(tool_name)
        if not handler:
            return {"error": f"Unknown tool: {tool_name}", "source": "sheldon"}

        try:
            return handler()
        except Exception as e:
            return {"error": str(e)[:500], "source": "sheldon"}

    # =========================================================================
    # Operations Tools (Snowflake/Redzone)
    # =========================================================================

    def _tool_plant_oee(self):
        data = self.snowflake.query(self.snowflake_queries['plant_oee'])
        return {"data": data, "source": "snowflake_redzone"}

    def _tool_oee_by_line(self):
        data = self.snowflake.query(self.snowflake_queries['oee_by_line'])
        return {"data": data, "source": "snowflake_redzone"}

    def _tool_red_flags(self):
        data = self.snowflake.query(self.snowflake_queries['lines_below_target'])
        return {"data": data, "source": "snowflake_redzone"}

    def _tool_top_downtime(self):
        data = self.snowflake.query(self.snowflake_queries['top_downtime'])
        return {"data": data, "source": "snowflake_redzone"}

    def _tool_labor_productivity(self):
        data = self.snowflake.query(self.snowflake_queries['labor_productivity'])
        return {"data": data, "source": "snowflake_redzone"}

    def _tool_oee_trend(self):
        data = self.snowflake.query(self.snowflake_queries['hourly_trend'])
        return {"data": data, "source": "snowflake_redzone"}

    def _tool_active_lines(self):
        data = self.snowflake.query(self.snowflake_queries['active_lines'])
        return {"data": data, "source": "snowflake_redzone"}

    # =========================================================================
    # Defect & Quality Analytics Tools (Snowflake)
    # =========================================================================

    def _tool_defect_summary(self, params):
        date_range = params.get("date_range", "today")
        julians = _julian_range(date_range)
        julian_list = ",".join(f"'{j}'" for j in julians)

        query = f"""
            SELECT "placeName" AS machine,
                   SUM(TRY_TO_NUMBER("values")) AS defect_count,
                   COUNT(DISTINCT "dataItemName") AS defect_types
            FROM ZGRZDCXCHH_DB."ameriqual-org"."v_completeddataitem"
            WHERE "runId" IN ({julian_list})
              AND "void" = 'False'
              AND "createdDate" >= DATEADD(year, -9, CURRENT_DATE())
              AND "dataSheetName" NOT LIKE '%Machine Reject%'
              AND "dataSheetName" NOT LIKE '%Changeover%'
              AND "dataSheetName" NOT LIKE '%Retort Load%'
              AND "dataSheetName" NOT LIKE '%Shift%'
              AND "dataSheetName" NOT LIKE '%Hourly%'
              AND "dataItemName" NOT IN (
                  'Total Number of Retort Pouches', 'Total Number of Defects',
                  'Code Time Frame', 'Is this the last code for the run?')
              AND TRY_TO_NUMBER("values") > 0
            GROUP BY "placeName"
            ORDER BY defect_count DESC
        """
        data = self.snowflake.query(query)
        return {"data": data, "source": "snowflake_redzone", "date_range": date_range, "julian_dates": julians}

    def _tool_defect_types(self, params):
        date_range = params.get("date_range", "today")
        julians = _julian_range(date_range)
        julian_list = ",".join(f"'{j}'" for j in julians)

        query = f"""
            SELECT "dataItemName" AS defect_type,
                   SUM(TRY_TO_NUMBER("values")) AS total_count,
                   COUNT(DISTINCT "placeName") AS machines_affected,
                   LISTAGG(DISTINCT "placeName", ', ') WITHIN GROUP (ORDER BY "placeName") AS machine_list
            FROM ZGRZDCXCHH_DB."ameriqual-org"."v_completeddataitem"
            WHERE "runId" IN ({julian_list})
              AND "void" = 'False'
              AND "createdDate" >= DATEADD(year, -9, CURRENT_DATE())
              AND "dataSheetName" NOT LIKE '%Machine Reject%'
              AND "dataSheetName" NOT LIKE '%Changeover%'
              AND "dataSheetName" NOT LIKE '%Retort Load%'
              AND "dataSheetName" NOT LIKE '%Shift%'
              AND "dataSheetName" NOT LIKE '%Hourly%'
              AND "dataItemName" NOT IN (
                  'Total Number of Retort Pouches', 'Total Number of Defects',
                  'Code Time Frame', 'Is this the last code for the run?')
              AND TRY_TO_NUMBER("values") > 0
            GROUP BY "dataItemName"
            HAVING SUM(TRY_TO_NUMBER("values")) > 0
            ORDER BY total_count DESC
            LIMIT 20
        """
        data = self.snowflake.query(query)
        return {"data": data, "source": "snowflake_redzone", "date_range": date_range}

    def _tool_shift_comparison(self):
        query = """
            SELECT "shiftName" AS shift,
                   ROUND(AVG(LEAST("oee", 100)), 1) AS avg_oee,
                   SUM("outCount") AS total_output,
                   ROUND(SUM("downSeconds")/3600, 1) AS downtime_hours,
                   COUNT(DISTINCT "locationName") AS lines
            FROM ZGRZDCXCHH_DB."ameriqual-org"."v_shift"
            WHERE "startTime" >= DATEADD(day, -1, CURRENT_TIMESTAMP())
            GROUP BY "shiftName"
            ORDER BY "shiftName"
        """
        data = self.snowflake.query(query)
        return {"data": data, "source": "snowflake_redzone"}

    # =========================================================================
    # Financial Tools (Sage X3)
    # =========================================================================

    def _tool_financial_snapshot(self):
        results = {}

        for key in ['mtd_revenue', 'ytd_revenue', 'gross_margin_mtd', 'cash_position', 'ar_aging']:
            try:
                sql = self.sql_queries.get(key, '')
                if sql:
                    data = self.sage.query(sql)
                    results[key] = data[0] if data and len(data) == 1 else data
            except Exception as e:
                results[f"{key}_error"] = str(e)[:200]

        # Calculate AR days from aging data
        try:
            ar = results.get('ar_aging', {})
            if isinstance(ar, dict):
                total_ar = float(ar.get('TotalAR', 0) or 0)
                if total_ar > 0:
                    current = float(ar.get('Current30', 0) or 0)
                    d31 = float(ar.get('Days31to60', 0) or 0)
                    d61 = float(ar.get('Days61to90', 0) or 0)
                    d90 = float(ar.get('Over90', 0) or 0)
                    results['ar_days'] = round(
                        ((current * 15) + (d31 * 45) + (d61 * 75) + (d90 * 120)) / total_ar, 1
                    )
        except Exception:
            pass

        return {"data": results, "source": "sage_x3"}

    def _tool_inventory(self):
        results = {}
        for key in ['inventory_value', 'finished_goods_value']:
            try:
                sql = self.sql_queries.get(key, '')
                if sql:
                    data = self.sage.query(sql)
                    results[key] = data
            except Exception as e:
                results[f"{key}_error"] = str(e)[:200]
        return {"data": results, "source": "sage_x3"}

    def _tool_top_customers(self):
        try:
            data = self.sage.query(self.sql_queries.get('top_customers', ''))
            return {"data": data, "source": "sage_x3"}
        except Exception as e:
            return {"error": str(e)[:200], "source": "sage_x3"}

    def _tool_ebitda(self):
        try:
            data = self.sage.query(self.sql_queries.get('ebitda_mtd', ''))
            return {"data": data, "source": "sage_x3"}
        except Exception as e:
            return {"error": str(e)[:200], "source": "sage_x3"}

    # =========================================================================
    # Quality Pipeline (Donna QA System — HTTP)
    # =========================================================================

    def _tool_quality_pipeline(self):
        """Get batch release pipeline from Donna API (localhost:5002)."""
        pipeline = {}

        endpoints = {
            "summary": "/api/preshipment/summary",
            "alerts": "/api/preshipment/alerts",
            "problems": "/api/preshipment/problems",
        }

        donna_available = False
        for key, path in endpoints.items():
            try:
                req = urllib.request.Request(
                    f"http://localhost:5002{path}",
                    headers={"Content-Type": "application/json"}
                )
                with urllib.request.urlopen(req, timeout=10) as resp:
                    pipeline[key] = json.loads(resp.read().decode())
                    donna_available = True
            except Exception as e:
                pipeline[f"{key}_error"] = str(e)[:100]

        if not donna_available:
            pipeline["note"] = (
                "Donna QA system is not currently running (localhost:5002). "
                "Quality pipeline data unavailable. Start Donna to enable batch release tracking."
            )

        return {"data": pipeline, "source": "donna_qa_system"}

    # =========================================================================
    # Jackie Deep Analytics (HTTP)
    # =========================================================================

    def _tool_ask_jackie(self, params):
        """Forward a question to Jackie's AI agent (localhost:5001)."""
        question = params.get("question", "")
        try:
            body = json.dumps({"query": question, "include_sql": False}).encode()
            req = urllib.request.Request(
                "http://localhost:5001/api/jackie",
                data=body,
                headers={"Content-Type": "application/json"}
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read().decode())
                return {
                    "data": {
                        "response": data.get("response", ""),
                        "queries_executed": data.get("queries_executed", []),
                        "data_sources": data.get("data_sources", [])
                    },
                    "source": "jackie_analytics"
                }
        except Exception as e:
            return {
                "error": f"Jackie is not currently running (localhost:5001): {str(e)[:100]}",
                "source": "jackie_analytics",
                "note": "Use the direct Snowflake tools (get_defect_summary, get_defect_types, get_oee_by_line, etc.) instead."
            }

    # =========================================================================
    # Calendar (MS Graph)
    # =========================================================================

    def _tool_calendar(self):
        try:
            today = datetime.now()
            end = today + timedelta(days=7)
            events = self.graph.get_user_calendar("dstraub@ameriqual.com", today, end)

            formatted = []
            for event in events:
                start_dt = event.get("start", {})
                end_dt = event.get("end", {})
                location = event.get("location", {})
                formatted.append({
                    "title": event.get("subject", "No Subject"),
                    "start": start_dt.get("dateTime", ""),
                    "end": end_dt.get("dateTime", ""),
                    "location": location.get("displayName", "") if isinstance(location, dict) else str(location),
                    "is_all_day": event.get("isAllDay", False),
                    "show_as": event.get("showAs", "busy")
                })

            if formatted:
                return {"data": formatted, "source": "microsoft_graph"}
            else:
                return {"data": [], "source": "microsoft_graph", "note": "No events found or Graph API unavailable."}

        except Exception as e:
            return {"error": str(e)[:200], "source": "microsoft_graph",
                    "note": "Calendar data unavailable. Graph API may not be configured."}

    # =========================================================================
    # Business Health Score
    # =========================================================================

    def _tool_health_score(self):
        """Calculate composite business health score from all available data."""
        weights = {
            'oee': 0.25, 'quality': 0.20, 'grossMargin': 0.20,
            'arDays': 0.15, 'inventory': 0.10, 'availability': 0.10
        }
        scores = {}

        # OEE Score
        try:
            plant_data = self.snowflake.query(self.snowflake_queries['plant_oee'])
            if plant_data:
                oee = float(plant_data[0].get('PLANT_OEE', 0) or 0)
                scores['oee'] = min(100, max(0, oee / 0.85 * 100))

                quality = float(plant_data[0].get('PLANT_QUALITY', 0) or 0)
                scores['quality'] = min(100, quality)

                avail = float(plant_data[0].get('PLANT_AVAILABILITY', 0) or 0)
                scores['availability'] = min(100, avail)
        except Exception:
            scores['oee'] = 50
            scores['quality'] = 50
            scores['availability'] = 50

        # Gross Margin Score
        try:
            margin_data = self.sage.query(self.sql_queries.get('gross_margin_mtd', ''))
            if margin_data:
                revenue = float(margin_data[0].get('LineRevenue', 0) or 0)
                cost = float(margin_data[0].get('LineCost', 0) or 0)
                if revenue > 0:
                    margin_pct = ((revenue - cost) / revenue) * 100
                    scores['grossMargin'] = min(100, max(0, margin_pct / 0.30 * 100))
        except Exception:
            scores['grossMargin'] = 50

        # AR Days Score
        try:
            ar_data = self.sage.query(self.sql_queries.get('ar_aging', ''))
            if ar_data:
                ar = ar_data[0]
                total_ar = float(ar.get('TotalAR', 0) or 0)
                if total_ar > 0:
                    current = float(ar.get('Current30', 0) or 0)
                    d31 = float(ar.get('Days31to60', 0) or 0)
                    d61 = float(ar.get('Days61to90', 0) or 0)
                    d90 = float(ar.get('Over90', 0) or 0)
                    ar_days = ((current * 15) + (d31 * 45) + (d61 * 75) + (d90 * 120)) / total_ar
                    if ar_days <= 30:
                        scores['arDays'] = 100
                    elif ar_days <= 45:
                        scores['arDays'] = 80
                    elif ar_days <= 60:
                        scores['arDays'] = 60
                    else:
                        scores['arDays'] = max(20, 100 - ar_days)
        except Exception:
            scores['arDays'] = 50

        # Inventory Score (default reasonable)
        scores.setdefault('inventory', 70)

        # Calculate weighted score
        total_score = 0
        total_weight = 0
        for metric, weight in weights.items():
            if metric in scores:
                total_score += scores[metric] * weight
                total_weight += weight

        final = round(total_score / total_weight) if total_weight > 0 else 0
        status = 'excellent' if final >= 80 else 'good' if final >= 65 else 'fair' if final >= 50 else 'needs_attention'

        return {
            "data": {
                "overall_score": final,
                "status": status,
                "components": {k: round(v, 1) for k, v in scores.items()},
                "weights": weights
            },
            "source": "sheldon_composite"
        }

    # =========================================================================
    # Departmental KPIs (Internal SQL Server)
    # =========================================================================

    def _tool_departmental_kpis(self, params):
        """Query departmental KPIs from internal SQL Server databases."""
        department = params.get('department', 'all')

        if not self.internal_db:
            return {"error": "Internal database client not configured", "source": "internal_sql"}

        results = {}
        errors = []

        # Filter queries by department
        queries_to_run = {}
        for name, qdef in self.dept_kpi_queries.items():
            if qdef.get('sql') is None:
                continue  # Skip manual-only KPIs
            if department.lower() != 'all' and qdef.get('department', '').lower() != department.lower():
                continue
            queries_to_run[name] = qdef

        if not queries_to_run:
            return {
                "data": {},
                "source": "internal_sql",
                "note": f"No automated KPI queries available for '{department}'. Some KPIs are manually tracked."
            }

        for name, qdef in queries_to_run.items():
            db_key = qdef['database']
            sql = qdef['sql']

            # Snowflake queries go through snowflake client
            if db_key == 'snowflake':
                try:
                    data = self.snowflake.query(sql)
                    results[name] = {
                        'department': qdef.get('department', 'Unknown'),
                        'data': data[:20] if data else [],
                        'row_count': len(data) if data else 0,
                        'notes': qdef.get('notes', '')
                    }
                except Exception as e:
                    errors.append(f"{name}: {str(e)[:100]}")
                continue

            # Internal DB queries
            try:
                data = self.internal_db.query(sql, database=db_key)
                results[name] = {
                    'department': qdef.get('department', 'Unknown'),
                    'data': data[:20] if data else [],
                    'row_count': len(data) if data else 0,
                    'notes': qdef.get('notes', '')
                }
            except Exception as e:
                errors.append(f"{name} (db={db_key}): {str(e)[:100]}")

        return {
            "data": results,
            "source": "internal_sql",
            "department_filter": department,
            "queries_attempted": len(queries_to_run),
            "queries_succeeded": len(results),
            "errors": errors if errors else None
        }

    # =========================================================================
    # S&OP Strategic Planning
    # =========================================================================

    def _tool_sop_status(self):
        """Read the latest S&OP report and return decision-ready snapshot."""
        return SOPReader.read_sop_snapshot()

    # =========================================================================
    # Unified Morning Brief
    # =========================================================================

    def generate_morning_brief(self) -> dict:
        """Generate a comprehensive Chief of Staff morning briefing.
        S&OP-anchored, action-driven, money-focused.
        """
        brief_prompt = """Generate Dennis Straub's morning briefing. You are his Chief of Staff — tell him exactly what to do today to make money and protect revenue.

## STEP 1: Pull ALL data (call these tools in parallel where possible)
- **get_sop_status** — THIS FIRST. The S&OP plan is the foundation of everything else.
- **get_plant_oee** + **get_red_flag_lines** + **get_top_downtime** — Current operations
- **get_financial_snapshot** + **get_inventory_status** — Money position
- **get_quality_pipeline** — Shipment blockers
- **get_departmental_kpis('all')** — All department performance vs targets
- **get_executive_calendar** — Dennis's day
- **get_business_health** — Composite score

## STEP 2: Synthesize into this EXACT format

### MONEY MOVES — Top 3 Things That Make or Lose You Money Today
For each item:
- **What's happening** (1 sentence, lead with the impact)
- **Dollar/unit impact** (quantify — estimate if needed, label estimates as such)
- **Action** (exactly what to do, who to call, by when)
- **S&OP connection** (how this affects the plan)

Rank by financial impact. Mix problems AND opportunities — don't just report bad news.

### PLAN vs REALITY — S&OP Execution Status
Show a compact table of active lines:
| Line | S&OP Volume | Days Over/Under | OEE Today | Status |
Where Status = On Track / At Risk / Constrained / Ahead

Below the table:
- Retort capacity utilization (% and cycles)
- Any lines where today's OEE threatens the monthly S&OP target
- Specific rebalancing recommendations if any line is constrained while another has slack

### FINANCIAL PULSE
- Revenue MTD vs pace to hit monthly target
- Cash position
- AR aging — flag anything crossing 60 days with customer name and amount
- Gross margin trend (expanding/compressing/flat)
Keep this to 3-4 bullets max.

### WATCH LIST
Quick-hit flags from quality, safety, warehouse, procurement — only items that need action within 48 hours. Skip anything that's on track. Include:
- Overdue QA releases (count + $ value at risk)
- Foreign matter incidents vs CAL
- Warehouse capacity if near limit
- Any KPI exceeding its CAL or missing target significantly

### WINS
1-2 things going well. Dennis needs to know what's working so he can reinforce it.

### BOTTOM LINE
One sentence. The single most important thing Dennis should focus on today and why.

## RULES
- NO generic advice. Every recommendation must reference specific lines, customers, people, or numbers.
- Estimates are OK but LABEL them: "~$50K estimated based on line rate"
- If S&OP data is unavailable, still frame operations against capacity and revenue impact
- Keep the entire briefing under 800 words — dense, no fluff
- Do NOT just list metrics. Connect → Conclude → Command."""

        return self.process_message(brief_prompt)

    def clear_history(self):
        """Clear conversation history."""
        self.conversation_history = []
